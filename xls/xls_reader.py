import abc
import os
import re
import sys
from threading import RLock

import django
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.abspath(os.path.join(BASE_DIR, os.pardir)))

os.environ['DJANGO_SETTINGS_MODULE'] = 'Inventarisation.settings'
django.setup()

from info.models import Server
from dictionary.models import Domain, ServerRoom, OS, IP, SoftwareCatalog, ServerFuture
from django.contrib.auth.models import User


class BaseManager(abc.ABC):
    cache = {}
    lock = RLock()

    def check_name(self, name_el):
        return True if name_el and name_el is not None else False

    def clean_name(self, el_name):
        return el_name.strip()

    def get_value(self, el_name, **kwargs):
        with self.lock:
            if not self.check_name(el_name):
                return None
            el_name = self.clean_name(el_name)
            value = self.cache.get(el_name, None)
            if value:
                return value
            value = self.get_from_db(el_name, **kwargs)
            if value:
                self.cache[el_name] = value
            return value

    @abc.abstractmethod
    def find_in_db(self, el_name):
        pass

    @abc.abstractmethod
    def create_in_db(self, el_name, *args, **kwargs):
        pass

    def get_from_db(self, el_name, **kwargs):
        value = self.find_in_db(el_name)
        if value:
            return value
        else:
            return self.create_in_db(el_name, **kwargs)


class RoomManager(BaseManager):
    matcher = (
        (re.compile(r'[Бб]ълижняя'), 'Ближняя'),
        (re.compile(r'[Лл]окальная'), 'Локальная'),
        (re.compile(r'[Дд]альняя[\s]*[НнHh]'), 'Дальняя H'),
        (re.compile(r'[Дд]альняя[\s]*[АфAa]'), 'Дальняя A'),
    )

    def clean_name(self, el_name):
        el_name = super().clean_name(el_name)
        for match, new_name in self.matcher:
            if match.match(el_name):
                return new_name
        return el_name

    def find_in_db(self, el_name):
        try:
            return ServerRoom.objects.get(name=el_name)
        except ServerRoom.DoesNotExist:
            return None

    def create_in_db(self, el_name, **kwargs):
        room = ServerRoom(name=el_name)
        room.save()
        return room


class DomainManager(BaseManager):

    def clean_name(self, el_name):
        return super().clean_name(el_name).lower()

    def find_in_db(self, el_name):
        try:
            return Domain.objects.get(name=el_name)
        except Domain.DoesNotExist:
            return None

    def create_in_db(self, el_name, **kwargs):
        domain = Domain(name=el_name)
        domain.save()
        return domain


class IPManager(BaseManager):

    def check_name(self, name_el):
        valid = super().check_name(name_el)
        if valid:
            try:
                import ipaddress
                ipaddress.ip_address(name_el)
                return True
            except ValueError:
                return False
        else:
            return False

    def find_in_db(self, el_name):
        try:
            return IP.objects.get(ip_address=el_name)
        except IP.DoesNotExist:
            return None

    def create_in_db(self, el_name, **kwargs):
        ip = IP(ip_address=el_name)
        ip.save()
        return ip


class SoftManager(BaseManager):

    def find_in_db(self, el_name):
        try:
            return SoftwareCatalog.objects.get(name=el_name)
        except SoftwareCatalog.DoesNotExist:
            return None

    def create_in_db(self, el_name, **kwargs):
        soft = SoftwareCatalog(name=el_name)
        soft.save()
        return soft


class FutureManager(BaseManager):

    def find_in_db(self, el_name, **kwargs):
        try:
            return ServerFuture.objects.get(name=el_name)
        except ServerFuture.DoesNotExist:
            return None

    def create_in_db(self, el_name, **kwargs):
        if 'description' in kwargs:
            soft = ServerFuture(name=el_name, display_name=kwargs['description'])
        else:
            soft = ServerFuture(name=el_name, **kwargs)
        soft.save()
        return soft


class OSManager(BaseManager):
    matcher = (
        (re.compile(r'.*[Ss]erver.*2012[\s]*R2.*'), 'Windows Server 2012 R2 Ver 6.3'),
        (re.compile(r'.*[Ss]erver.*2012.*'), 'Windows Server 2012 Ver 6.2'),
        (re.compile(r'.*[Ss]erver[\s]*2016[\s]*[Dd]ata[Cc]enter.*'), 'Windows Server DataCenter 2016'),
        (re.compile(r'.*[Ss]erver[\s]*2016.*'), 'Windows Server Standard 2016'),
        (re.compile(r'.*[Ss]erver[\s]*2019[\s]*[Dd]ata[Cc]enter.*'), 'Windows Server DataCenter 2019'),
        (re.compile(r'.*[Ss]erver[\s]*2019.*'), 'Windows Server Standard 2019'),
        (re.compile(r'.*[Ss]erver[\s]*2022[\s]*[Dd]ata[Cc]enter.*'), 'Windows Server DataCenter 2022'),
        (re.compile(r'.*[Ss]erver[\s]*2022.*'), 'Windows Server Standard 2022'),
        (re.compile(r'.*[Ww]indows[\s]*10.*[Ee]nt.*'), 'Windows 10 Enterprise'),
        (re.compile(r'.*[Ww]indows[\s]*10.*[Pp]ro.*'), 'Windows 10 Pro'),
        (re.compile(r'.*[Ww]indows[\s]*11.*[Ee]nt.*'), 'Windows 11 Enterprise'),
        (re.compile(r'.*[Ww]indows[\s]*11.*[Pp]ro.*'), 'Windows 11 Pro'),
        (re.compile(r'.*2008.*'), 'Windows Server Enterprise 2008 R2 SP1'),
        (re.compile(r'.*2003.*'), 'Windows Server 2003'),
    )

    def clean_name(self, el_name):
        el_name = super().clean_name(el_name)
        for match, new_name in self.matcher:
            if match.match(el_name):
                return new_name
        return el_name

    def find_in_db(self, el_name):
        try:
            return OS.objects.get(name=el_name)
        except OS.DoesNotExist:
            return None

    def create_in_db(self, el_name, **kwargs):
        os = OS(name=el_name)
        os.save()
        return os


if __name__ == "__main__":
    rooms = RoomManager()
    domains = DomainManager()
    ipsm = IPManager()
    softm = SoftManager()
    osm = OSManager()
    admin_user = User.objects.get(pk=1)
    softs_re = re.compile(r'([\r\n]|[\s]{2,})')

    wb = load_workbook(r'\\bs.local.erc\IT\BSPD\IT infrastructure\ERC_BS_Servers_2022_1_3.xlsx')
    print(wb.sheetnames)
    ws = wb['SERVERS']
    for x in range(2, 84):
        room = rooms.get_value(ws.cell(row=x, column=1).value)
        if not room:
            print(f"ROW {x} NO ROOM NAME")
            continue
        domain = domains.get_value(ws.cell(row=x, column=3).value)
        if not domain:
            print(f"ROW {x} NO DOMAIN NAME")
            continue
        ip_list = ws.cell(row=x, column=2).value
        ips = [ipsm.get_value(ip) for ip in ip_list.split('\n')]

        name = ws.cell(row=x, column=4).value

        if not name:
            print(f"ROW {x} NO NAME")
            continue
        else:
            name = name.split("\n")[0]
        spft_list = ws.cell(row=x, column=12).value
        if spft_list:
            list_installed_soft = softs_re.split(spft_list)
            list_installed_soft = filter(lambda a: True if len(a.strip()) >= 1 else False, list_installed_soft)
            softs = [softm.get_value(soft.strip()) for soft in list_installed_soft]
        else:
            softs = []

        os_info = ws.cell(row=x, column=6).value
        if not os_info:
            print(f"ROW {x} NO OS")
            continue
        os_desc = os_info.split('\n')
        os = osm.get_value(os_desc[0])

        ver = os_desc[1] if len(os_desc) > 1 else None

        zabix = True if ws.cell(row=x, column=8).value == '+' else False
        comments = ws.cell(row=x, column=14).value
        virtual_machine = ws.cell(row=x, column=5).value
        virtual_machine = virtual_machine.strip() if virtual_machine else None
        server = Server(name=name.upper(),
                        domain=domain,
                        room=room,
                        os_name=os,
                        virtual_server_name=virtual_machine,
                        os_version=ver,
                        description=comments,
                        has_monitoring_agent=zabix,
                        updated_by=admin_user,
                        )
        server.save()
        server.ip_addresses.set(ips)
        server.installed_soft.set(softs)
        server.save()
